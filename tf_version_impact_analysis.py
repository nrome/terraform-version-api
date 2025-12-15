import os
import re
import json
import tempfile
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import subprocess
import platform


# ------------------------------------------------------------------------------
# Extract rc.type values from all policy files in /policies
# ------------------------------------------------------------------------------
def extract_rc_types_from_policies(policies_dir="policies"):
    results = []

    if not os.path.isdir(policies_dir):
        raise Exception(f"Directory not found: {policies_dir}")

    for root, _, files in os.walk(policies_dir):
        for file in files:
            file_path = os.path.join(root, file)

            try:
                with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                    content = f.read()
            except Exception:
                print(f"Warning: couldn't read file {file_path}")
                continue

            # Regex: rc.type is "some_name"
            matches = re.findall(r'rc\.type\s+is\s+"([^"]+)"', content)

            for m in matches:
                results.append({
                    "filename": file,
                    "resource_type": m.strip()
                })

    return results


# ------------------------------------------------------------------------------
# Load local Terraform Registry JSON
# Expected format:
#   { "resources": [ { "type": "azurerm_virtual_network", "title": "..." }, ... ] }
# ------------------------------------------------------------------------------
def load_local_registry(registry_path):
    with open(registry_path, "r") as f:
        return json.load(f)


# ------------------------------------------------------------------------------
# Compare rc.type to registry titles
# ------------------------------------------------------------------------------
def compare_types(resource_entries, registry_data):

    registry_map = {entry["type"]: entry["title"] for entry in registry_data.get("resources", [])}

    results = []

    for entry in resource_entries:
        rtype = entry["resource_type"]
        expected_title = registry_map.get(rtype)

        mismatch = expected_title is None

        results.append({
            "filename": entry["filename"],
            "resource_type": rtype,
            "registry_title": expected_title if expected_title else "❌ No match",
            "match": "Match" if not mismatch else "Mismatch"
        })

    return results


# ------------------------------------------------------------------------------
# Write Excel with color highlighting
# ------------------------------------------------------------------------------
def write_excel(results):

    df = pd.DataFrame(results)
    df = df.sort_values(["filename", "resource_type"])

    temp_dir = tempfile.gettempdir()
    excel_path = os.path.join(temp_dir, "terraform_policy_type_comparison.xlsx")

    # Write the base file
    df.to_excel(excel_path, index=False)

    # Apply color highlighting
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    match_col = None
    for idx, col in enumerate(ws[1], 1):
        if col.value == "match":
            match_col = idx
            break

    if match_col:
        for row in ws.iter_rows(min_row=2):
            cell = row[match_col - 1]

            if cell.value == "Match":
                for c in row:
                    c.fill = green_fill
            else:
                for c in row:
                    c.fill = red_fill

    wb.save(excel_path)

    # Auto-launch the file
    system = platform.system()
    try:
        if system == "Windows":
            os.startfile(excel_path)
        elif system == "Darwin":
            subprocess.run(["open", excel_path])
        else:
            subprocess.run(["xdg-open", excel_path])
    except Exception:
        print(f"Excel created but could not auto-open: {excel_path}")

    print(f"\nExcel file created at: {excel_path}")
    return excel_path


# ------------------------------------------------------------------------------
# JSON export for API / CLI use
# ------------------------------------------------------------------------------
def export_json(results):
    return json.dumps(results, indent=2)


# ------------------------------------------------------------------------------
# MAIN
# ------------------------------------------------------------------------------
def main():
    registry_path = "terraform_registry.json"  # Local registry file

    resource_entries = extract_rc_types_from_policies("policies")
    registry_data = load_local_registry(registry_path)

    results = compare_types(resource_entries, registry_data)

    excel_path = write_excel(results)

    print("\nJSON Export:\n")
    print(export_json(results))


if __name__ == "__main__":
    main()
